from pathlib import Path

from openpyxl import load_workbook

from music_calendar_finder.db import connect_db, init_db, upsert_city
from music_calendar_finder.exporter import export_master
from music_calendar_finder.pipeline import RunOptions, process_city
from music_calendar_finder.reporting import overall_status, report


CONFIG = {
    "budgets": {
        "small": {
            "top_sources_target": 25,
            "verified_sources_target": 25,
            "max_verified_sources": 50,
            "max_candidate_domains": 100,
            "max_serpapi_queries": 3,
            "max_query_pages": 1,
        },
        "music_market_overrides": [],
        "tourism_market_overrides": [],
    },
    "serpapi": {"num": 10},
    "crawl": {"max_pages_per_domain": 4},
    "dedupe": {"low_priority_domains": []},
}


def test_pipeline_dry_run_excluded_from_live_export_and_reporting(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "Austin", "state": "TX"})
        city = conn.execute("select * from cities where id=?", (city_id,)).fetchone()
        result = process_city(conn, city, CONFIG, RunOptions(dry_run=True, mode="test"))
        output = export_master(conn, tmp_path / "exports")
        status = overall_status(conn)
        data = report(conn, top=10)
    workbook = load_workbook(output, read_only=True)
    city_summary_headers = [cell.value for cell in next(workbook["City Summary"].iter_rows(min_row=1, max_row=1))]
    top_source_headers = [cell.value for cell in next(workbook["Top Sources"].iter_rows(min_row=1, max_row=1))]
    assert result.status == "completed"
    assert "City Summary" in workbook.sheetnames
    assert "Top Sources" in workbook.sheetnames
    assert "population" in city_summary_headers
    assert "city_tier" in city_summary_headers
    assert "population" in top_source_headers
    assert "city_tier" in top_source_headers
    assert status["completed"] == 0
    assert workbook["Verified Sources"].max_row == 1
    assert workbook["Top Sources"].max_row == 1
    assert data["top_cities"][0]["verified_sources"] == 0
