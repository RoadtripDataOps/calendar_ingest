from __future__ import annotations

from openpyxl import load_workbook

from music_calendar_finder.db import connect_db, init_db, upsert_city
from music_calendar_finder.exporter import build_export_frames, export_master
from music_calendar_finder.quality import classify_url_quality


BAD_NEEDLES = (
    "google.com/calendar/event",
    "google.com/calendar/render",
    "calendar.google.com/calendar/embed",
    "calendar.google.com/calendar/render",
    "axs.com/events/",
    "open.spotify.com/concerts/location",
    "concertfix.com",
    "real-estate",
    "realestate",
)


def test_phase6_url_quality_classifies_known_bad_patterns():
    assert classify_url_quality("https://calendar.google.com/calendar/render?cid=abc") == "google_calendar_embed"
    assert classify_url_quality("https://calendar.google.com/calendar/u/0/r/eventedit?text=Show") == "google_calendar_event_template"
    assert classify_url_quality("https://www.axs.com/events/12345/some-show-tickets") == "individual_ticket_page"
    assert classify_url_quality("https://open.spotify.com/concerts/location/los-angeles") == "national_aggregator"
    assert classify_url_quality("https://concertfix.com/concerts/austin-tx") == "national_aggregator"
    assert classify_url_quality("https://austinrealestatedarsh.com/blog/live-music") == "weak_article_or_blog_post"


def test_phase6_export_sanitizes_all_url_fields_and_reconciles_summary_counts(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "Austin", "state": "TX", "population": 974447, "city_tier": "large"})
        run_id = _insert_completed_run(conn, city_id)
        _insert_source(
            conn,
            city_id,
            run_id,
            "austinchronicle.com",
            "Austin Chronicle",
            "alt_weekly",
            "https://www.austinchronicle.com/",
            "https://calendar.google.com/calendar/render?cid=austin",
            music_url="https://www.austinchronicle.com/music/",
            score=94,
        )
        _insert_source(
            conn,
            city_id,
            run_id,
            "austinrealestatedarsh.com",
            "Austin Real Estate Darsh Music Guide",
            "local_publication",
            "https://austinrealestatedarsh.com/",
            "https://austinrealestatedarsh.com/blog/austin-live-music/",
            score=99,
        )
        _insert_source(
            conn,
            city_id,
            run_id,
            "axs.com",
            "AXS Austin Tickets",
            "ticketing_platform",
            "https://www.axs.com/events/12345/show-tickets",
            "https://www.axs.com/events/12345/show-tickets",
            score=98,
        )
        _insert_source(
            conn,
            city_id,
            run_id,
            "concertfix.com",
            "ConcertFix Austin",
            "event_aggregator",
            "https://concertfix.com/concerts/austin-tx",
            "https://concertfix.com/concerts/austin-tx",
            score=97,
        )
        _insert_source(
            conn,
            city_id,
            run_id,
            "spotify.com",
            "Spotify Austin Concerts",
            "event_aggregator",
            "https://open.spotify.com/concerts/location/austin",
            "https://open.spotify.com/concerts/location/austin",
            score=96,
        )
        _insert_candidate(conn, city_id, run_id, "https://example.org/pending-events")

        frames = build_export_frames(conn)

    top = frames["Top Sources"]
    verified = frames["Verified Sources"]
    summary = frames["City Summary"].iloc[0].to_dict()
    top_text = _frame_text(top)
    verified_text = _frame_text(verified)

    assert "website_url_quality" in top.columns
    assert "best_calendar_url_quality" in top.columns
    assert "source_quality_tier" in top.columns
    assert len(top) == 1
    assert top.iloc[0]["source_name"] == "Austin Chronicle"
    assert not top.iloc[0]["best_calendar_url"]
    assert top.iloc[0]["best_calendar_url_quality"] == ""
    for needle in BAD_NEEDLES:
        assert needle not in top_text.casefold()
        assert needle not in verified_text.casefold()
    assert summary["raw_candidates_count"] == 1
    assert summary["candidates_unverified_count"] == len(frames["Candidates Unverified"])
    assert summary["top_sources_count"] == len(top)
    assert summary["verified_sources_count"] == len(verified)
    assert summary["search_api_queries_count"] == 0


def test_phase6_export_caps_normal_sheet_cells_below_10000_chars(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "Austin", "state": "TX", "population": 974447, "city_tier": "large"})
        run_id = _insert_completed_run(conn, city_id)
        _insert_source(
            conn,
            city_id,
            run_id,
            "do512.com",
            "Do512 Austin Events",
            "event_aggregator",
            "https://do512.com/",
            "https://do512.com/events",
            social_links='["' + ("https://do512.com/events/very-long-link," * 800) + '"]',
            score=95,
        )
        output = export_master(conn, tmp_path / "exports")

    workbook = load_workbook(output, read_only=True, data_only=True)
    for sheet_name in ["Top Sources", "Verified Sources", "City Summary", "Candidates Unverified", "Rejected"]:
        for row in workbook[sheet_name].iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str):
                    assert len(value) <= 10000


def test_phase6_export_strips_excel_illegal_control_characters(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "New Orleans", "state": "LA", "population": 1029300, "city_tier": "mega"})
        run_id = _insert_completed_run(conn, city_id)
        _insert_source(
            conn,
            city_id,
            run_id,
            "artsneworleans.org",
            "Arts\x03 District New Orleans",
            "local_calendar",
            "https://artsneworleans.org/",
            "https://artsneworleans.org/events",
            score=95,
        )
        output = export_master(conn, tmp_path / "exports")

    workbook = load_workbook(output, read_only=True, data_only=True)
    values = [
        value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    ]
    assert all("\x03" not in value for value in values)
    assert any("Arts District New Orleans" in value for value in values)


def _insert_completed_run(conn, city_id: int) -> int:
    conn.execute(
        "insert into city_runs (city_id, run_mode, status, mode, queries_planned, queries_completed) values (?, 'live', 'completed', 'test', 1, 1)",
        (city_id,),
    )
    conn.commit()
    return int(conn.execute("select last_insert_rowid()").fetchone()[0])


def _insert_source(
    conn,
    city_id: int,
    run_id: int,
    root_domain: str,
    source_name: str,
    source_category: str,
    website_url: str,
    best_calendar_url: str,
    *,
    music_url: str | None = None,
    social_links: str = "[]",
    score: float = 80,
) -> None:
    conn.execute(
        """
        insert into source_pages (
          city_id, run_id, run_mode, root_domain, url_origin, source_name,
          source_category, website_url, best_calendar_url, music_url,
          social_links_json, title, meta_description, local_relevance_score,
          music_signal_score, calendar_signal_score, authority_score,
          freshness_score, diversity_score, total_score, confidence,
          status, url_validation_status
        )
        values (?, ?, 'live', ?, 'serpapi_organic', ?, ?, ?, ?, ?, ?, ?, ?, 80, 80, 80, 70, 60, 80, ?, 'high', 'verified', 'valid')
        """,
        (
            city_id,
            run_id,
            root_domain,
            source_name,
            source_category,
            website_url,
            best_calendar_url,
            music_url,
            social_links,
            f"{source_name} Austin music calendar",
            "Austin music, arts, culture, events, calendar, and live music.",
            score,
        ),
    )
    conn.commit()


def _insert_candidate(conn, city_id: int, run_id: int, url: str) -> None:
    conn.execute(
        """
        insert into candidate_urls (
          city_id, run_id, source_url, normalized_url, domain, root_domain,
          run_mode, url_origin, url_validation_status, status
        )
        values (?, ?, ?, ?, 'example.org', 'example.org', 'live', 'serpapi_organic', 'timeout', 'candidate_unverified')
        """,
        (city_id, run_id, url, url),
    )
    conn.commit()


def _frame_text(frame) -> str:
    return "\n".join(str(value) for value in frame.astype(str).to_numpy().flatten())
