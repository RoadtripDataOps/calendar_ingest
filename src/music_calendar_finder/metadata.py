from __future__ import annotations

from pathlib import Path
import sqlite3
from typing import Any

from openpyxl import load_workbook

from .utils import city_tier_for_population, normalize_blank, normalize_key, to_float, to_int, utc_now


def hydrate_missing_city_metadata(conn: sqlite3.Connection, workbook_path: str | Path) -> int:
    path = Path(workbook_path)
    if not path.exists():
        fallback = path.parent / "uscities.xlsx"
        if fallback.exists():
            path = fallback
        else:
            return 0
    rows = conn.execute(
        """
        select id, city, state, population, city_tier
        from cities
        where population is null
           or city_tier is null
           or city_tier='tiny'
        """
    ).fetchall()
    if not rows:
        return 0
    wanted = {(normalize_key(row["city"]), normalize_key(row["state"])): row for row in rows}
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: index for index, header in enumerate(headers)}
    required = {"city", "state_id", "population"}
    if not required.issubset(header_index):
        return 0
    hydrated = 0
    for values in sheet.iter_rows(min_row=2, values_only=True):
        city = normalize_blank(_value(values, header_index, "city"))
        state = normalize_blank(_value(values, header_index, "state_id"))
        if not city or not state:
            continue
        key = (normalize_key(city), normalize_key(state))
        city_row = wanted.get(key)
        if not city_row:
            continue
        population = to_int(_value(values, header_index, "population"))
        conn.execute(
            """
            update cities
            set state_name=coalesce(state_name, ?),
                population=coalesce(population, ?),
                lat=coalesce(lat, ?),
                lng=coalesce(lng, ?),
                original_row_id=coalesce(original_row_id, ?),
                city_tier=?,
                updated_at=?
            where id=?
            """,
            (
                normalize_blank(_value(values, header_index, "state_name")),
                population,
                to_float(_value(values, header_index, "lat")),
                to_float(_value(values, header_index, "lng")),
                str(_value(values, header_index, "id")) if _value(values, header_index, "id") is not None else None,
                city_tier_for_population(population),
                utc_now(),
                city_row["id"],
            ),
        )
        hydrated += 1
    conn.commit()
    return hydrated


def hydrate_city_metadata(conn: sqlite3.Connection, city_id: int, workbook_path: str | Path) -> sqlite3.Row:
    hydrate_missing_city_metadata(conn, workbook_path)
    return conn.execute("select * from cities where id=?", (city_id,)).fetchone()


def _value(values: tuple[Any, ...], header_index: dict[str, int], column: str) -> Any:
    index = header_index.get(column)
    if index is None or index >= len(values):
        return None
    return values[index]

