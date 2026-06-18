from __future__ import annotations

from collections import Counter

from openpyxl import Workbook, load_workbook

from music_calendar_finder.db import connect_db, init_db, record_provider_usage, save_rejected_source, upsert_city
from music_calendar_finder.exporter import build_export_frames, export_master
from music_calendar_finder.maintenance import clear_live_city_data
from music_calendar_finder.metadata import hydrate_missing_city_metadata
from music_calendar_finder.models import SourceCandidate
from music_calendar_finder.score import score_source
from music_calendar_finder.validation import source_url_rejection_reason


def test_export_defaults_to_latest_completed_live_run_and_keeps_history_separate(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(
            conn,
            {"city": "Austin", "state": "TX", "population": 974447, "city_tier": "large"},
        )
        old_run = _insert_run(conn, city_id)
        latest_run = _insert_run(conn, city_id)
        _insert_source(conn, city_id, old_run, "old-austin.example", score=92)
        _insert_source(conn, city_id, latest_run, "new-austin.example", score=88)
        _insert_source(conn, city_id, latest_run, "eventbrite.com", score=99)
        _insert_candidate(conn, city_id, old_run, "https://old-austin.example/events")
        _insert_candidate(conn, city_id, latest_run, "https://new-austin.example/events")
        save_rejected_source(conn, city_id, old_run, "old-reject.example", "https://old-reject.example/event/1", "old")
        save_rejected_source(conn, city_id, latest_run, "new-reject.example", "https://new-reject.example/event/1", "new")
        record_provider_usage(conn, "serpapi", old_run, city_id, "search", "old query")
        record_provider_usage(conn, "serpapi", latest_run, city_id, "search", "new query")

        output = export_master(conn, tmp_path / "exports")
        history_output = export_master(conn, tmp_path / "exports-history", include_history=True)

    workbook = load_workbook(output, read_only=True)
    live_text = "\n".join(_sheet_text(workbook, name) for name in workbook.sheetnames)
    assert "new-austin.example" in live_text
    assert "new query" in live_text
    assert "eventbrite.com" not in live_text
    assert "old-austin.example" not in live_text
    assert "old query" not in live_text
    assert "Historical Sources" not in workbook.sheetnames

    summary = _sheet_records(workbook, "City Summary")[0]
    assert summary["top_sources_count"] == 1
    assert summary["verified_sources_count"] == 1
    assert summary["raw_candidates_count"] == 1
    assert summary["candidates_unverified_count"] == 1
    assert summary["rejected_count"] == 1

    history_workbook = load_workbook(history_output, read_only=True)
    assert "Run History" in history_workbook.sheetnames
    assert "Historical Sources" in history_workbook.sheetnames
    assert "Historical Rejections" in history_workbook.sheetnames
    assert "Historical Search API Usage" in history_workbook.sheetnames
    assert "old-austin.example" in _sheet_text(history_workbook, "Historical Sources")
    assert "old query" in _sheet_text(history_workbook, "Historical Search API Usage")


def test_los_angeles_top_sources_count_matches_exported_top_rows(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(
            conn,
            {"city": "Los Angeles", "state": "CA", "population": 3898747, "city_tier": "mega"},
        )
        run_id = _insert_run(conn, city_id)
        for index in range(60):
            _insert_source(conn, city_id, run_id, f"los-angeles-source-{index}.example", score=100 - index)
        _insert_source(conn, city_id, run_id, "nytimes.com", score=100)

        frames = build_export_frames(conn)

    city_summary = frames["City Summary"]
    top_sources = frames["Top Sources"]
    assert int(city_summary.loc[0, "top_sources_count"]) == 50
    assert len(top_sources[top_sources["city"] == "Los Angeles"]) == 50
    assert int(city_summary.loc[0, "verified_sources_count"]) == 60


def test_force_live_cleanup_removes_only_live_city_data(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "Austin", "state": "TX"})
        live_run = _insert_run(conn, city_id, run_mode="live")
        dry_run = _insert_run(conn, city_id, run_mode="dry_run")
        _insert_source(conn, city_id, live_run, "live.example", run_mode="live")
        _insert_source(conn, city_id, dry_run, "dry.example", run_mode="dry_run")
        _insert_candidate(conn, city_id, live_run, "https://live.example/events", run_mode="live")
        _insert_candidate(conn, city_id, dry_run, "https://dry.example/events", run_mode="dry_run")
        save_rejected_source(conn, city_id, live_run, "live-reject.example", "https://live-reject.example", "bad")
        save_rejected_source(
            conn,
            city_id,
            dry_run,
            "dry-reject.example",
            "https://dry-reject.example",
            "bad",
            run_mode="dry_run",
        )
        record_provider_usage(conn, "serpapi", live_run, city_id, "search", "live", run_mode="live")
        record_provider_usage(conn, "serpapi", dry_run, city_id, "search", "dry", run_mode="dry_run")

        removed = clear_live_city_data(conn, city_id)

        assert removed["city_runs"] == 1
        assert _count(conn, "city_runs", "run_mode='live'") == 0
        assert _count(conn, "source_pages", "run_mode='live'") == 0
        assert _count(conn, "candidate_urls", "run_mode='live'") == 0
        assert _count(conn, "rejected_sources", "run_mode='live'") == 0
        assert _count(conn, "provider_usage", "run_mode='live'") == 0
        assert _count(conn, "city_runs", "run_mode='dry_run'") == 1
        assert _count(conn, "source_pages", "run_mode='dry_run'") == 1


def test_metadata_hydration_backfills_population_tier_and_coordinates(tmp_path):
    db_path = tmp_path / "test.sqlite"
    workbook_path = tmp_path / "uscities.xlsx"
    _write_city_metadata_workbook(workbook_path)
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "Austin", "state": "TX"})
        hydrated = hydrate_missing_city_metadata(conn, workbook_path)
        row = conn.execute("select * from cities where id=?", (city_id,)).fetchone()

    assert hydrated == 1
    assert row["population"] == 974447
    assert row["city_tier"] == "large"
    assert row["lat"] == 30.2672
    assert row["lng"] == -97.7431
    assert row["original_row_id"] == "12345"


def test_quality_hardening_rejects_or_demotes_weak_source_urls():
    assert source_url_rejection_reason("https://seatgeek.com/nashville-concert-tickets") == "ticketing_or_national_aggregator"
    assert source_url_rejection_reason("https://www.nytimes.com/section/arts/music") == "ticketing_or_national_aggregator"
    assert source_url_rejection_reason("https://calendar.google.com/calendar/event?eid=abc") == "google_calendar_event_template"
    assert source_url_rejection_reason("https://example.com/affiliate/redirect?url=https://tickets.example") == "affiliate_or_redirect_url"
    assert source_url_rejection_reason("https://example.com/austin-real-estate/music-scene") == "weak_article_or_blog_post"

    local_source = _score_source(
        root_domain="austinchronicle.com",
        website_url="https://www.austinchronicle.com/",
        best_calendar_url="https://www.austinchronicle.com/events/music/",
        source_category="alt_weekly",
    )
    national_source = _score_source(
        root_domain="nytimes.com",
        website_url="https://www.nytimes.com/",
        best_calendar_url="https://www.nytimes.com/events/los-angeles-music.html",
        source_category="newspaper",
        city="Los Angeles",
        state="CA",
    )
    ticket_source = _score_source(
        root_domain="seatgeek.com",
        website_url="https://seatgeek.com/",
        best_calendar_url="https://seatgeek.com/austin-concert-tickets",
        source_category="ticketing_platform",
    )

    assert local_source.total_score >= 50
    assert national_source.total_score < 50
    assert ticket_source.total_score < 50


def _insert_run(conn, city_id: int, *, run_mode: str = "live") -> int:
    conn.execute(
        """
        insert into city_runs (city_id, run_mode, status, mode, queries_planned, queries_completed)
        values (?, ?, 'completed', 'test', 1, 1)
        """,
        (city_id, run_mode),
    )
    conn.commit()
    return int(conn.execute("select last_insert_rowid()").fetchone()[0])


def _insert_source(
    conn,
    city_id: int,
    run_id: int,
    root_domain: str,
    *,
    score: float = 80,
    run_mode: str = "live",
) -> None:
    conn.execute(
        """
        insert into source_pages (
          city_id, run_id, run_mode, root_domain, url_origin, source_name,
          source_category, website_url, best_calendar_url, local_relevance_score,
          music_signal_score, calendar_signal_score, total_score, confidence,
          status, url_validation_status
        )
        values (?, ?, ?, ?, 'serpapi_organic', ?, 'local_publication', ?, ?, 80, 80, 80, ?, 'high', 'verified', 'valid')
        """,
        (
            city_id,
            run_id,
            run_mode,
            root_domain,
            root_domain,
            f"https://{root_domain}/",
            f"https://{root_domain}/events",
            score,
        ),
    )
    conn.commit()


def _insert_candidate(conn, city_id: int, run_id: int, url: str, *, run_mode: str = "live") -> None:
    root = url.split("//", 1)[1].split("/", 1)[0]
    conn.execute(
        """
        insert into candidate_urls (
          city_id, run_id, source_url, normalized_url, domain, root_domain,
          run_mode, url_origin, url_validation_status, status
        )
        values (?, ?, ?, ?, ?, ?, ?, 'serpapi_organic', 'timeout', 'candidate_unverified')
        """,
        (city_id, run_id, url, url, root, root, run_mode),
    )
    conn.commit()


def _count(conn, table: str, where: str) -> int:
    return int(conn.execute(f"select count(*) from {table} where {where}").fetchone()[0])


def _write_city_metadata_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["city", "state_id", "state_name", "lat", "lng", "population", "id"])
    sheet.append(["Austin", "TX", "Texas", 30.2672, -97.7431, 974447, 12345])
    workbook.save(path)


def _score_source(
    *,
    root_domain: str,
    website_url: str,
    best_calendar_url: str,
    source_category: str,
    city: str = "Austin",
    state: str = "TX",
):
    source = SourceCandidate(
        city_id=1,
        root_domain=root_domain,
        website_url=website_url,
        best_calendar_url=best_calendar_url,
        source_category=source_category,
        title=f"{city} music events calendar",
        meta_description=f"{city} live music, concerts, arts, festivals, and things to do.",
        detected_keywords=["events", "music", "calendar"],
        body_text=f"{city} {state} upcoming events live music calendar submit event 2026.",
    )
    return score_source(source, {"city": city, "state": state}, Counter())


def _sheet_records(workbook, sheet_name: str) -> list[dict[str, object]]:
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]]


def _sheet_text(workbook, sheet_name: str) -> str:
    values = []
    for row in workbook[sheet_name].iter_rows(values_only=True):
        values.extend(str(cell) for cell in row if cell is not None)
    return "\n".join(values)
