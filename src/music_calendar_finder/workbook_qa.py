from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .quality import is_bad_source_name


@dataclass
class WorkbookQaIssue:
    severity: str
    sheet: str
    city: str
    message: str
    source_name: str | None = None


TOP_SOURCE_TIER_CAPS = {
    "venue_calendar": 0.40,
    "regional_reference": 0.20,
}

BAD_TITLE_TERMS = (
    "checking your browser",
    "just a moment",
)

COUNT_COLUMNS = {
    "Top Sources": "top_sources_count",
    "Verified Sources": "verified_sources_count",
    "Candidates Unverified": "candidates_unverified_count",
    "Rejected": "rejected_count",
    "Search API Usage": "search_api_queries_count",
}

HOMEPAGE_ONLY_ALLOWED_TIERS = {
    "arts_culture_org",
    "core_local_calendar",
    "core_local_publication",
    "festival_hub",
    "music_org",
    "promoter_calendar",
    "radio_media",
    "tourism_cvb_calendar",
    "university_calendar",
    "venue_calendar",
}


def qa_workbook(path: str | Path) -> list[WorkbookQaIssue]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    issues: list[WorkbookQaIssue] = []
    if "Top Sources" not in workbook.sheetnames:
        return [WorkbookQaIssue("error", "Workbook", "", "Missing Top Sources sheet")]
    top_rows = _records(workbook, "Top Sources")
    summary_rows = _records(workbook, "City Summary") if "City Summary" in workbook.sheetnames else []
    issues.extend(_qa_top_sources(top_rows))
    issues.extend(_qa_count_reconciliation(workbook, summary_rows))
    return issues


def _qa_top_sources(rows: list[dict[str, Any]]) -> list[WorkbookQaIssue]:
    issues: list[WorkbookQaIssue] = []
    by_city: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        city = str(row.get("city") or "")
        state = str(row.get("state") or "")
        source_name = str(row.get("source_name") or "")
        title = str(row.get("title") or "")
        tier = str(row.get("source_quality_tier") or "")
        by_city[(city, state)].append(row)
        haystack = f"{source_name} {title}".casefold()
        if any(term in haystack for term in BAD_TITLE_TERMS):
            issues.append(WorkbookQaIssue("error", "Top Sources", _city_label(city, state), "Challenge/interstitial title in Top Sources", source_name))
        if is_bad_source_name(source_name) or is_bad_source_name(title):
            issues.append(WorkbookQaIssue("error", "Top Sources", _city_label(city, state), "Generic or unusable source name/title in Top Sources", source_name))
        if tier == "weak_article_or_blog_post":
            issues.append(WorkbookQaIssue("error", "Top Sources", _city_label(city, state), "Weak article/blog post in Top Sources", source_name))
        if tier in {"ticketing_platform", "national_aggregator", "individual_event_page", "real_estate_or_commercial_blog"}:
            issues.append(WorkbookQaIssue("error", "Top Sources", _city_label(city, state), f"Disallowed tier {tier} in Top Sources", source_name))
        if not row.get("best_calendar_url_quality") and row.get("website_url_quality") == "homepage_only" and tier not in HOMEPAGE_ONLY_ALLOWED_TIERS:
            issues.append(WorkbookQaIssue("error", "Top Sources", _city_label(city, state), "Homepage-only source without acceptable tier", source_name))
    for (city, state), city_rows in by_city.items():
        limit = 50 if city in {"Los Angeles", "New York"} else 25
        counts = Counter(str(row.get("source_quality_tier") or "") for row in city_rows)
        for tier, ratio in TOP_SOURCE_TIER_CAPS.items():
            cap = int(limit * ratio)
            if counts[tier] > cap:
                issues.append(
                    WorkbookQaIssue(
                        "error",
                        "Top Sources",
                        _city_label(city, state),
                        f"{tier} count {counts[tier]} exceeds cap {cap}",
                    )
                )
    return issues


def _qa_count_reconciliation(workbook, summary_rows: list[dict[str, Any]]) -> list[WorkbookQaIssue]:
    issues: list[WorkbookQaIssue] = []
    if not summary_rows:
        return issues
    for sheet_name, summary_column in COUNT_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            issues.append(WorkbookQaIssue("error", sheet_name, "", f"Missing sheet {sheet_name}"))
            continue
        counts = Counter((row.get("city"), row.get("state")) for row in _records(workbook, sheet_name))
        for row in summary_rows:
            key = (row.get("city"), row.get("state"))
            expected = int(row.get(summary_column) or 0)
            actual = counts.get(key, 0)
            if expected != actual:
                issues.append(
                    WorkbookQaIssue(
                        "error",
                        "City Summary",
                        _city_label(str(key[0] or ""), str(key[1] or "")),
                        f"{summary_column}={expected} but {sheet_name} has {actual}",
                    )
                )
    return issues


def _records(workbook, sheet_name: str) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        return []
    return [dict(zip(headers, row)) for row in rows]


def _city_label(city: str, state: str) -> str:
    return f"{city}, {state}".strip(", ")
