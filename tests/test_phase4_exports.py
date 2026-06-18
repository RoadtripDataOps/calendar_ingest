from openpyxl import load_workbook
from typer.testing import CliRunner

from music_calendar_finder.cli import app
from music_calendar_finder.db import connect_db, create_city_run, finish_city_run, init_db, save_candidate_url, upsert_city
from music_calendar_finder.exporter import export_master
from music_calendar_finder.maintenance import clean_dry_run_data, validate_city_urls
from music_calendar_finder.validation import UrlValidator


BAD_AUSTIN_URLS = [
    "https://austin-tx-altarts.org/calendar",
    "https://austin-tx-altmusic.org/events",
    "https://austin-tx-chamberarts.org/calendar",
    "https://austin-tx-chambermusic.org/events",
    "https://austin-tx-concertsarts.org/calendar",
    "https://austin-tx-concertsmusic.org/events",
    "https://austin-tx-newspaperarts.org/calendar",
    "https://austin-tx-newspapermusic.org/events",
    "https://visitaustin-tx-alt.com/events",
    "https://visitaustin-tx-chamber.com/events",
    "https://visitaustin-tx-newspaper.com/events",
    "https://visitaustin-tx-tourism.com/events",
]


def test_bad_austin_urls_rejected_and_excluded_from_verified_sheets(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "Austin", "state": "TX"})
        run_id = create_city_run(conn, city_id, "test")
        for url in BAD_AUSTIN_URLS:
            save_candidate_url(
                conn,
                {
                    "city_id": city_id,
                    "run_id": run_id,
                    "source_url": url,
                    "normalized_url": url,
                    "root_domain": url.split("//", 1)[1].split("/", 1)[0],
                    "domain": url.split("//", 1)[1].split("/", 1)[0],
                    "run_mode": "live",
                    "url_origin": "serpapi_organic",
                },
            )
        finish_city_run(conn, run_id, "completed")
        city = conn.execute("select * from cities where id=?", (city_id,)).fetchone()
        validate_city_urls(conn, city, validator=UrlValidator())
        output = export_master(conn, tmp_path / "exports")
    workbook = load_workbook(output, read_only=True)
    verified_values = _sheet_text(workbook, "Verified Sources")
    top_values = _sheet_text(workbook, "Top Sources")
    rejected_values = _sheet_text(workbook, "Rejected")
    for url in BAD_AUSTIN_URLS:
        assert url not in verified_values
        assert url not in top_values
        assert url in rejected_values
    assert "source_url" in [cell.value for cell in next(workbook["Rejected"].iter_rows(min_row=1, max_row=1))]
    assert "rejection_reason" in [cell.value for cell in next(workbook["Rejected"].iter_rows(min_row=1, max_row=1))]


def test_clean_dry_run_data_removes_old_and_new_markers(tmp_path):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        city_id, _ = upsert_city(conn, {"city": "Austin", "state": "TX"})
        conn.execute("insert into city_runs (city_id, run_mode, status) values (?, 'dry_run', 'completed')", (city_id,))
        run_id = conn.execute("select last_insert_rowid()").fetchone()[0]
        conn.execute("insert into search_queries (run_id, city_id, run_mode, query, api_status) values (?, ?, 'dry_run', 'q', 'dry_run')", (run_id, city_id))
        save_candidate_url(
            conn,
            {
                "city_id": city_id,
                "run_id": run_id,
                "source_url": "https://dry-run.example/events",
                "normalized_url": "https://dry-run.example/events",
                "run_mode": "dry_run",
                "url_origin": "dry_run_fixture",
            },
        )
        conn.commit()
        result = clean_dry_run_data(conn)
        remaining = conn.execute("select count(*) from candidate_urls").fetchone()[0]
    assert result.candidate_urls == 1
    assert result.search_queries == 1
    assert result.city_runs == 1
    assert remaining == 0


def test_validate_urls_cli_command_is_wired(tmp_path, monkeypatch):
    db_path = tmp_path / "test.sqlite"
    init_db(db_path)
    with connect_db(db_path) as conn:
        upsert_city(conn, {"city": "Austin", "state": "TX"})

    class FakeResult:
        candidates_checked = 1
        sources_checked = 0
        rejected = 1
        verified_kept = 0
        unverified = 0

        @property
        def __dict__(self):
            return {
                "candidates_checked": self.candidates_checked,
                "sources_checked": self.sources_checked,
                "rejected": self.rejected,
                "verified_kept": self.verified_kept,
                "unverified": self.unverified,
            }

    monkeypatch.setattr("music_calendar_finder.cli.validate_city_urls", lambda conn, row: FakeResult())
    result = CliRunner().invoke(app, ["validate-urls", "--city", "Austin", "--state", "TX", "--database-path", str(db_path)])
    assert result.exit_code == 0
    assert "Validate URLs" in result.output


def _sheet_text(workbook, sheet_name: str) -> str:
    values = []
    for row in workbook[sheet_name].iter_rows(values_only=True):
        values.extend(str(cell) for cell in row if cell is not None)
    return "\n".join(values)
