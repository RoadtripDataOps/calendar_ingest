from __future__ import annotations

from pathlib import Path
import re
import sqlite3
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .quality import (
    EXPORTED_URL_QUALITY_FIELDS,
    URL_QUALITY_FIELDS,
    classify_url_quality,
    is_top_source_eligible,
    is_verified_source_eligible,
    sanitize_url_fields,
    source_tier_sort_key,
)
from .utils import ensure_dir, slugify, utc_now


MASTER_FILENAME = "master_music_calendar_sources.xlsx"
EXPORT_CELL_CHAR_LIMIT = 9500
EXCEL_ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010\013\014\016-\037]")
TOP_SOURCE_TIER_CAPS = {
    "venue_calendar": 0.40,
    "regional_reference": 0.20,
}
BASE_SOURCE_FILTER = """
  sp.status='verified'
  and coalesce(sp.run_mode, 'live')='live'
  and sp.url_validation_status in ('valid', 'redirect_valid', 'forbidden_but_real')
  and sp.url_origin in ('serpapi_organic', 'serpapi_sitelink', 'crawled_internal_link', 'canonical_url', 'redirect_final_url', 'manual_seed')
"""


def export_master(
    conn: sqlite3.Connection,
    export_dir: str | Path,
    *,
    write_city_workbooks: bool = False,
    run_mode: str = "live",
    include_history: bool = False,
) -> Path:
    export_dir = ensure_dir(Path(export_dir) / "dry_run") if run_mode == "dry_run" else ensure_dir(export_dir)
    output_path = export_dir / MASTER_FILENAME
    sheets = build_export_frames(conn, run_mode=run_mode, include_history=include_history)
    _write_workbook(output_path, sheets)
    conn.execute(
        "insert into exports (export_path, export_type, run_mode, created_at, row_count) values (?, 'master', ?, ?, ?)",
        (str(output_path), run_mode, utc_now(), len(sheets["Verified Sources"])),
    )
    conn.commit()
    if write_city_workbooks:
        export_city_workbooks(conn, export_dir, run_mode=run_mode, include_history=include_history)
    return output_path


def export_city_workbooks(
    conn: sqlite3.Connection,
    export_dir: str | Path,
    *,
    run_mode: str = "live",
    include_history: bool = False,
) -> list[Path]:
    city_dir = ensure_dir(Path(export_dir) / "cities")
    output: list[Path] = []
    cities = conn.execute("select id, city, state from cities order by state, city").fetchall()
    for city in cities:
        sheets = build_export_frames(conn, city_id=city["id"], run_mode=run_mode, include_history=include_history)
        filename = f"{slugify(city['state'] or 'na')}_{slugify(city['city'])}_music_calendar_sources.xlsx"
        path = city_dir / filename
        _write_workbook(path, sheets)
        conn.execute(
            "insert into exports (export_path, export_type, run_mode, city_id, created_at, row_count) values (?, 'city', ?, ?, ?, ?)",
            (str(path), run_mode, city["id"], utc_now(), len(sheets["Verified Sources"])),
        )
        output.append(path)
    conn.commit()
    return output


def build_export_frames(
    conn: sqlite3.Connection,
    city_id: int | None = None,
    *,
    run_mode: str = "live",
    include_history: bool = False,
) -> dict[str, pd.DataFrame]:
    params = _scoped_params(run_mode, city_id)
    city_clause = "and c.id=?" if city_id else ""
    city_base = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select c.id as city_id, lr.run_id, c.city, c.state, c.country, c.metro_name,
               c.population, c.city_tier, c.last_processed_at, c.status,
               c.error_message as notes
        from latest_runs lr
        join cities c on c.id=lr.city_id
        where 1=1 {city_clause}
        order by c.state, c.city
        """,
        conn,
        params=params,
    )

    raw_sources = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select c.id as city_id, c.city, c.state, c.country, c.metro_name, c.population, c.city_tier,
               sp.root_domain, sp.source_name, sp.source_category,
               sp.website_url, sp.best_calendar_url, sp.music_url, sp.events_url, sp.arts_url,
               sp.tourism_url, sp.about_url, sp.rss_url, sp.contact_email,
               sp.social_links_json as social_links, sp.title, sp.meta_description,
               sp.detected_keywords_json as detected_keywords, sp.local_relevance_score,
               sp.music_signal_score, sp.calendar_signal_score, sp.authority_score,
               sp.freshness_score, sp.diversity_score, sp.total_score, sp.confidence,
               sp.status, sp.why_selected, sp.robots_allowed, sp.crawl_status,
               sp.error_message, sp.url_validation_status, sp.http_status, sp.final_url,
               sp.resolved_domain, sp.content_type, sp.page_title, sp.validation_error,
               sp.first_seen_at, sp.last_checked_at
        from source_pages sp
        join latest_runs lr on lr.city_id=sp.city_id and lr.run_id=sp.run_id
        join cities c on c.id=sp.city_id
        where {BASE_SOURCE_FILTER}
        {city_clause}
        order by c.city, c.state, sp.total_score desc
        """,
        conn,
        params=params,
    )
    sources = _prepare_source_frame(raw_sources)
    verified = _verified_sources_frame(sources)
    top_sources = _top_sources_frame(sources)

    raw_candidate_counts = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select cu.city_id, count(*) as raw_candidates_count
        from candidate_urls cu
        join latest_runs lr on lr.city_id=cu.city_id and lr.run_id=cu.run_id
        join cities c on c.id=cu.city_id
        where 1=1 {city_clause}
        group by cu.city_id
        """,
        conn,
        params=params,
    )

    search_api_usage = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select c.id as city_id, pu.provider, pu.run_id, c.city, c.state, c.population, c.city_tier,
               pu.endpoint, pu.query, pu.used_cache, pu.created_at
        from provider_usage pu
        join latest_runs lr on lr.city_id=pu.city_id and lr.run_id=pu.run_id
        left join cities c on c.id=pu.city_id
        where 1=1 {city_clause}
        order by pu.created_at desc
        """,
        conn,
        params=params,
    )

    candidates_unverified = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select c.id as city_id, c.city, c.state, c.country, c.population, c.city_tier,
               cu.source_url, cu.normalized_url, cu.domain,
               cu.root_domain, cu.title, cu.snippet, cu.provider, sq.query, cu.position,
               cu.status, cu.url_origin, cu.url_validation_status, cu.http_status,
               cu.final_url, cu.resolved_domain, cu.content_type, cu.page_title,
               cu.validation_error, cu.validated_at, cu.first_seen_at, cu.last_seen_at
        from candidate_urls cu
        join latest_runs lr on lr.city_id=cu.city_id and lr.run_id=cu.run_id
        join cities c on c.id=cu.city_id
        left join search_queries sq on sq.id=cu.query_id
        where (?='dry_run' or coalesce(cu.url_origin, '') != 'dry_run_fixture')
          and (
            cu.url_validation_status is null
            or cu.url_validation_status in ('timeout', 'forbidden_but_real', 'robots_disallowed')
            or cu.status in ('low_priority', 'candidate_unverified')
          )
          {city_clause}
        order by c.city, c.state, cu.root_domain, cu.position
        """,
        conn,
        params=_history_params(run_mode, city_id),
    )
    candidates_unverified = _add_url_quality_columns(candidates_unverified, ("source_url", "normalized_url", "final_url"))

    rejected = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select c.id as city_id, c.city, c.state, c.country, c.population, c.city_tier,
               rs.url as source_url, rs.root_domain as domain, rs.reason as rejection_reason,
               rs.details, rs.url_validation_status, rs.http_status, rs.validation_error,
               rs.url_origin, rs.final_url, rs.resolved_domain, rs.content_type,
               rs.page_title, rs.validated_at, rs.created_at
        from rejected_sources rs
        join latest_runs lr on lr.city_id=rs.city_id and lr.run_id=rs.run_id
        join cities c on c.id=rs.city_id
        where 1=1 {city_clause}
        order by c.city, c.state, rs.created_at
        """,
        conn,
        params=params,
    )
    rejected = _add_url_quality_columns(rejected, ("source_url", "final_url"))

    city_summary = _city_summary_frame(
        city_base,
        raw_candidate_counts,
        verified,
        top_sources,
        candidates_unverified,
        rejected,
        search_api_usage,
    )

    sheets = {
        "City Summary": city_summary,
        "Top Sources": _drop_internal_columns(top_sources),
        "Verified Sources": _drop_internal_columns(verified),
        "Candidates Unverified": _drop_internal_columns(candidates_unverified),
        "Rejected": _drop_internal_columns(rejected),
        "Search API Usage": _drop_internal_columns(search_api_usage),
    }
    if include_history:
        sheets.update(_history_frames(conn, run_mode=run_mode, city_id=city_id))
    return sheets


def _prepare_source_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        for field in EXPORTED_URL_QUALITY_FIELDS:
            frame[f"{field}_quality"] = []
        frame["source_quality_tier"] = []
        return frame
    rows = [sanitize_url_fields(row) for row in frame.to_dict(orient="records")]
    prepared = pd.DataFrame(rows)
    for field in EXPORTED_URL_QUALITY_FIELDS:
        column = f"{field}_quality"
        if column not in prepared:
            prepared[column] = ""
    prepared["source_quality_tier"] = prepared.apply(lambda row: row.get("source_quality_tier") or "", axis=1)
    if "content_quality_status" not in prepared:
        prepared["content_quality_status"] = "unknown"
    if "challenge_page_detected" not in prepared:
        prepared["challenge_page_detected"] = 0
    return prepared


def _verified_sources_frame(sources: pd.DataFrame) -> pd.DataFrame:
    if sources.empty:
        return _source_columns(sources.copy(), include_rank=False)
    verified = sources[sources.apply(lambda row: is_verified_source_eligible(row), axis=1)].copy()
    verified["_source_tier_rank"] = verified["source_quality_tier"].map(source_tier_sort_key)
    verified = verified.sort_values(
        ["city", "state", "_source_tier_rank", "total_score", "source_name"],
        ascending=[True, True, True, False, True],
    )
    return _source_columns(verified, include_rank=False)


def _top_sources_frame(sources: pd.DataFrame) -> pd.DataFrame:
    if sources.empty:
        return _source_columns(sources.copy(), include_rank=True)
    top = sources[sources.apply(lambda row: is_top_source_eligible(row), axis=1)].copy()
    if top.empty:
        return _source_columns(top, include_rank=True)
    top["_source_tier_rank"] = top["source_quality_tier"].map(source_tier_sort_key)
    top = top.sort_values(
        ["city", "state", "_source_tier_rank", "total_score", "source_name"],
        ascending=[True, True, True, False, True],
    )
    top["rank"] = top.groupby(["city", "state"]).cumcount() + 1
    top["top_limit"] = top.apply(lambda row: 50 if row.get("city") in {"Los Angeles", "New York"} else 25, axis=1)
    top = _apply_top_source_caps(top)
    return _source_columns(top, include_rank=True)


def _apply_top_source_caps(top: pd.DataFrame) -> pd.DataFrame:
    if top.empty:
        return top
    selected_frames: list[pd.DataFrame] = []
    for _, group in top.groupby(["city", "state"], sort=False):
        top_limit = int(group["top_limit"].iloc[0] or 25)
        caps = {tier: int(top_limit * ratio) for tier, ratio in TOP_SOURCE_TIER_CAPS.items()}
        counts: dict[str, int] = {}
        rows = []
        for _, row in group.iterrows():
            if len(rows) >= top_limit:
                break
            tier = str(row.get("source_quality_tier") or "")
            cap = caps.get(tier)
            if cap is not None and counts.get(tier, 0) >= cap:
                continue
            rows.append(row)
            counts[tier] = counts.get(tier, 0) + 1
        if rows:
            city_frame = pd.DataFrame(rows)
            city_frame["rank"] = range(1, len(city_frame) + 1)
            selected_frames.append(city_frame)
    if not selected_frames:
        return top.iloc[0:0].copy()
    return pd.concat(selected_frames, ignore_index=True)


def _source_columns(frame: pd.DataFrame, *, include_rank: bool) -> pd.DataFrame:
    quality_columns = [f"{field}_quality" for field in EXPORTED_URL_QUALITY_FIELDS]
    base_columns = [
        "city_id",
        "city",
        "state",
        "country",
        "metro_name",
        "population",
        "city_tier",
        "source_quality_tier",
        "content_quality_status",
        "challenge_page_detected",
    ]
    if include_rank:
        base_columns.append("rank")
    base_columns.extend(
        [
            "source_name",
            "source_category",
            "website_url",
            "best_calendar_url",
            "music_url",
            "events_url",
            "arts_url",
            "tourism_url",
            "about_url",
            "rss_url",
            *quality_columns,
            "contact_email",
            "social_links",
            "title",
            "meta_description",
            "detected_keywords",
            "local_relevance_score",
            "music_signal_score",
            "calendar_signal_score",
            "authority_score",
            "freshness_score",
            "diversity_score",
            "total_score",
            "confidence",
            "status",
            "why_selected",
            "robots_allowed",
            "crawl_status",
            "error_message",
            "url_validation_status",
            "http_status",
            "final_url",
            "resolved_domain",
            "content_type",
            "page_title",
            "validation_error",
            "first_seen_at",
            "last_checked_at",
        ]
    )
    return _select_columns(frame, base_columns)


def _city_summary_frame(
    city_base: pd.DataFrame,
    raw_candidate_counts: pd.DataFrame,
    verified: pd.DataFrame,
    top_sources: pd.DataFrame,
    candidates_unverified: pd.DataFrame,
    rejected: pd.DataFrame,
    search_api_usage: pd.DataFrame,
) -> pd.DataFrame:
    summary = city_base.copy()
    if summary.empty:
        return _select_columns(
            summary,
            [
                "city",
                "state",
                "country",
                "metro_name",
                "population",
                "city_tier",
                "last_processed_at",
                "status",
                "raw_candidates_count",
                "verified_sources_count",
                "candidates_unverified_count",
                "rejected_count",
                "top_sources_count",
                "search_api_queries_count",
                "avg_music_signal_score",
                "avg_calendar_signal_score",
                "coverage_grade",
                "notes",
            ],
        )

    count_maps = {
        "raw_candidates_count": _count_map(raw_candidate_counts, "raw_candidates_count"),
        "verified_sources_count": _frame_count_map(verified),
        "candidates_unverified_count": _frame_count_map(candidates_unverified),
        "rejected_count": _frame_count_map(rejected),
        "top_sources_count": _frame_count_map(top_sources),
        "search_api_queries_count": _frame_count_map(search_api_usage),
    }
    for column, values in count_maps.items():
        summary[column] = summary["city_id"].map(values).fillna(0).astype(int)

    avg_scores = verified.groupby("city_id")[["music_signal_score", "calendar_signal_score"]].mean() if not verified.empty else pd.DataFrame()
    summary["avg_music_signal_score"] = summary["city_id"].map(avg_scores["music_signal_score"].round(1).to_dict() if not avg_scores.empty else {}).fillna(0)
    summary["avg_calendar_signal_score"] = summary["city_id"].map(avg_scores["calendar_signal_score"].round(1).to_dict() if not avg_scores.empty else {}).fillna(0)
    summary["coverage_grade"] = summary["verified_sources_count"].map(_coverage_grade)
    return _select_columns(
        summary,
        [
            "city",
            "state",
            "country",
            "metro_name",
            "population",
            "city_tier",
            "last_processed_at",
            "status",
            "raw_candidates_count",
            "verified_sources_count",
            "candidates_unverified_count",
            "rejected_count",
            "top_sources_count",
            "search_api_queries_count",
            "avg_music_signal_score",
            "avg_calendar_signal_score",
            "coverage_grade",
            "notes",
        ],
    )


def _add_url_quality_columns(frame: pd.DataFrame, fields: tuple[str, ...]) -> pd.DataFrame:
    if frame.empty:
        for field in fields:
            frame[f"{field}_quality"] = []
        return frame
    output = frame.copy()
    for field in fields:
        if field not in output:
            continue
        output[f"{field}_quality"] = output.apply(
            lambda row, name=field: classify_url_quality(
                row.get(name),
                source_root_domain=row.get("root_domain") or row.get("domain"),
                field_name=name,
            ),
            axis=1,
        )
    return output


def _count_map(frame: pd.DataFrame, count_column: str) -> dict[int, int]:
    if frame.empty:
        return {}
    return {int(row["city_id"]): int(row[count_column]) for row in frame.to_dict(orient="records")}


def _frame_count_map(frame: pd.DataFrame) -> dict[int, int]:
    if frame.empty or "city_id" not in frame:
        return {}
    return frame.groupby("city_id").size().astype(int).to_dict()


def _coverage_grade(count: int) -> str:
    if count >= 50:
        return "A"
    if count >= 25:
        return "B"
    if count >= 10:
        return "C"
    if count > 0:
        return "D"
    return "F"


def _select_columns(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    output = frame.copy()
    for column in columns:
        if column not in output:
            output[column] = None
    return output[columns]


def _drop_internal_columns(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.drop(columns=[column for column in ("city_id", "run_id", "_source_tier_rank", "top_limit") if column in frame], errors="ignore")


def _history_frames(conn: sqlite3.Connection, *, run_mode: str, city_id: int | None) -> dict[str, pd.DataFrame]:
    city_clause = "and c.id=?" if city_id else ""
    params = _history_params(run_mode, city_id)
    run_history = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select cr.id as run_id, c.city, c.state, cr.run_mode, cr.started_at, cr.finished_at,
               cr.status, cr.mode, cr.queries_planned, cr.queries_completed,
               cr.candidates_found, cr.verified_sources, cr.rejected_sources,
               case when lr.run_id=cr.id then 1 else 0 end as is_latest,
               cr.error_message
        from city_runs cr
        join cities c on c.id=cr.city_id
        left join latest_runs lr on lr.city_id=cr.city_id
        where coalesce(cr.run_mode, 'live')=? {city_clause}
        order by c.state, c.city, cr.id desc
        """,
        conn,
        params=params,
    )
    historical_sources = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select c.city, c.state, sp.run_id, sp.source_name, sp.source_category,
               sp.website_url, sp.best_calendar_url, sp.total_score, sp.status,
               sp.url_validation_status, sp.url_origin, sp.last_checked_at
        from source_pages sp
        join cities c on c.id=sp.city_id
        left join latest_runs lr on lr.city_id=sp.city_id
        where coalesce(sp.run_mode, 'live')=? and (lr.run_id is null or sp.run_id != lr.run_id)
        {city_clause}
        order by c.state, c.city, sp.run_id desc, sp.total_score desc
        """,
        conn,
        params=params,
    )
    historical_rejections = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select c.city, c.state, rs.run_id, rs.url as source_url, rs.root_domain,
               rs.reason, rs.url_validation_status, rs.validation_error, rs.created_at
        from rejected_sources rs
        join cities c on c.id=rs.city_id
        left join latest_runs lr on lr.city_id=rs.city_id
        where coalesce(rs.run_mode, 'live')=? and (lr.run_id is null or rs.run_id != lr.run_id)
        {city_clause}
        order by c.state, c.city, rs.run_id desc, rs.created_at
        """,
        conn,
        params=params,
    )
    historical_usage = pd.read_sql_query(
        f"""
        with latest_runs as ({_latest_runs_sql()})
        select pu.provider, pu.run_id, c.city, c.state, pu.endpoint, pu.query, pu.used_cache, pu.created_at
        from provider_usage pu
        join cities c on c.id=pu.city_id
        left join latest_runs lr on lr.city_id=pu.city_id
        where coalesce(pu.run_mode, 'live')=? and (lr.run_id is null or pu.run_id != lr.run_id)
        {city_clause}
        order by c.state, c.city, pu.run_id desc, pu.created_at
        """,
        conn,
        params=params,
    )
    return {
        "Run History": run_history,
        "Historical Sources": historical_sources,
        "Historical Rejections": historical_rejections,
        "Historical Search API Usage": historical_usage,
    }


def _latest_runs_sql() -> str:
    return """
      select cr.city_id, max(cr.id) as run_id
      from city_runs cr
      where coalesce(cr.run_mode, 'live')=? and cr.status='completed'
      group by cr.city_id
    """


def _scoped_params(run_mode: str, city_id: int | None = None) -> tuple[Any, ...]:
    return (run_mode, city_id) if city_id else (run_mode,)


def _history_params(run_mode: str, city_id: int | None = None) -> tuple[Any, ...]:
    return (run_mode, run_mode, city_id) if city_id else (run_mode, run_mode)


def _write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    ensure_dir(path.parent)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            _truncate_frame(frame).to_excel(writer, sheet_name=name[:31], index=False)
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            header = str(column_cells[0].value or "")
            width = min(max(12, len(header) + 2), 44)
            for cell in column_cells[1:50]:
                if cell.value is not None:
                    width = min(max(width, len(str(cell.value)) + 2), 60)
                if header.endswith("_score") or header == "total_score":
                    cell.number_format = "0.0"
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def _truncate_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    suffix = " ... [truncated]"

    def clean_cell(value: Any) -> Any:
        if isinstance(value, str):
            value = EXCEL_ILLEGAL_CHARACTERS_RE.sub("", value)
            if len(value) > EXPORT_CELL_CHAR_LIMIT:
                return value[: EXPORT_CELL_CHAR_LIMIT - len(suffix)] + suffix
        return value

    return frame.map(clean_cell)
